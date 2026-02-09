from fastapi import APIRouter, HTTPException, UploadFile, File, Form, Depends
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from app.schemas.company import CompanyInput, CredibilityAnalysis
from app.schemas.allocation import AllocationRequest, AllocationResponse
from app.engine.pipeline_orchestrator import PipelineOrchestrator
from app.engine.factory import get_ai_provider
from app.engine.allocation_engine import AllocationEngine
from app.core.document_parser import DocumentParser
from app.core.database import get_db
from openpyxl import load_workbook
import logging
import shutil
import os

router = APIRouter(prefix="/verification", tags=["Verification"])
logger = logging.getLogger(__name__)

@router.post("/verify", response_model=CredibilityAnalysis)
async def verify_company(data: CompanyInput, db: AsyncSession = Depends(get_db)):
    """verifies company legitimacy via registry, footprint, and ai analysis."""
    try:
        orchestrator = PipelineOrchestrator()
        result = await orchestrator.run_pipeline(data, db)
        return result
    except Exception as e:
        logger.error(f"verification failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/parse/recruiter-registration")
async def parse_recruiter_registration(file: UploadFile = File(...)):
    """parses recruiter registration doc and returns structured data."""
    temp_path = f"outputs/temp_{file.filename}"
    os.makedirs("outputs", exist_ok=True)
    
    try:
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        raw = DocumentParser.parse(temp_path)
        ai = get_ai_provider()
        extracted_data = ai.extract_company_input(raw['content'])
        
        if extracted_data.get("error"):
             raise HTTPException(status_code=400, detail=extracted_data["error"])
             
        return extracted_data
        
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"parsing failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)

@router.post("/parse/offer-letter")
async def parse_offer_letter(
    file: UploadFile = File(None),
    student_programme: str = Form(None),
    offer_text: str = Form(None)
):
    """
    parses offer letter and checks relevance to student's programme.
    
    accepts either:
    - file upload (pdf/docx) + student_programme
    - offer_text (raw text) + student_programme
    
    student_programme examples:
    - "bachelor of science in economics, statistics and mathematics"
    - "btech computer science and engineering"
    - "msc data science"
    """
    logger.info(f"offer-letter endpoint hit - file: {file.filename if file else 'none'}, programme: {student_programme}")
    
    if not file and not offer_text:
        raise HTTPException(status_code=400, detail="must provide either 'file' or 'offer_text'")
    
    programme_context = student_programme if student_programme else "not specified"
    
    temp_path = None
    try:
        if file:
            temp_path = f"outputs/temp_{file.filename}"
            os.makedirs("outputs", exist_ok=True)
            with open(temp_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            raw = DocumentParser.parse(temp_path)
            content = raw.get('content', '')
        else:
            content = offer_text
        
        logger.info(f"content length: {len(content)} chars, programme: {programme_context}")
        
        ai = get_ai_provider()
        extracted_data = ai.extract_offer_details(content)
        relevance = ai.verify_internship_relevance(content, programme_context)
        
        logger.info(f"extraction complete - is_relevant: {relevance.get('is_relevant', 'n/a')}")
        
        return {
            "extracted_data": extracted_data,
            "relevance_analysis": relevance,
            "student_programme": programme_context
        }
        
    except HTTPException as he:
        raise he
    except Exception as e:
        logger.error(f"parsing failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)

@router.get("/report/{filename}")
async def get_report(filename: str):
    """downloads pdf report."""
    file_path = f"reports/{filename}"
    if os.path.exists(file_path):
        return FileResponse(file_path, media_type="application/pdf", filename=filename)
    return {"error": "file not found"}

@router.post("/allocation/recommend", response_model=AllocationResponse)
async def recommend_guide(request: AllocationRequest):
    """recommends faculty guide based on expertise match."""
    engine = AllocationEngine()
    result = engine.allocate(request)
    return result

@router.post("/allocation/validate-pair")
async def validate_allocation_pair(request: dict):
    """validates manual student-faculty pairing."""
    engine = AllocationEngine()
    student = request.get("student")
    faculty = request.get("faculty")
    
    if not student or not faculty:
        raise HTTPException(status_code=400, detail="missing data")
        
    return engine.validate_pair(student, faculty)

@router.get("/history")
async def get_verification_history():
    """returns verification history from master excel log."""
    log_path = "reports/master_log.xlsx"
    if not os.path.exists(log_path):
        return {"history": []}
        
    try:
        wb = load_workbook(log_path, read_only=True)
        ws = wb.active
        
        data = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            entry = dict(zip(headers, row))
            data.append(entry)
            
        return {"count": len(data), "history": data}
    except Exception as e:
        logger.error(f"failed to read history: {e}")
        raise HTTPException(status_code=500, detail="could not read history log")
