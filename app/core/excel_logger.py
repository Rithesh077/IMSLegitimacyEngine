import os
from datetime import datetime
import logging
import openpyxl
from openpyxl import Workbook, load_workbook
from app.schemas.company import CredibilityAnalysis, CompanyInput

logger = logging.getLogger(__name__)

class ExcelLogger:
    FILE_PATH = "reports/master_log.xlsx"
    
    HEADERS = [
        "Timestamp", "Company Name", "Trust Score", "Tier", "Status",
        "Registry Found", "HR Verified", "Address Verified", "Email Domain Match",
        "LinkedIn Match", "Website Match", "Report Path"
    ]

    @classmethod
    def log_verification(cls, input_data: CompanyInput, analysis: CredibilityAnalysis):
        try:
            # Ensure directory exists
            os.makedirs("reports", exist_ok=True)
            
            # Load or Create Workbook
            if os.path.exists(cls.FILE_PATH):
                wb = load_workbook(cls.FILE_PATH)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(cls.HEADERS)
            
            # Prepare Data Row
            signals = analysis.details.get("signals", {})
            row = [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                input_data.name,
                analysis.trust_score,
                analysis.trust_tier,
                analysis.verification_status,
                signals.get("registry_link_found", False),
                signals.get("hr_verified", False),
                signals.get("address_verified", False),
                signals.get("email_domain_match", False),
                signals.get("linkedin_verified", False),
                signals.get("website_content_match", False),
                analysis.details.get("report_path", "N/A")
            ]
            
            ws.append(row)
            wb.save(cls.FILE_PATH)
            logger.info(f"Logged to Excel: {cls.FILE_PATH}")
            
        except Exception as e:
            logger.error(f"Failed to log to Excel: {e}")
